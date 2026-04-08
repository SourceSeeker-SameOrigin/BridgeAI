import csv
import io

from app.rag.parsers.base import BaseParser


class ExcelParser(BaseParser):
    """Parse Excel (.xlsx, .xls) and CSV (.csv) files to plain text."""

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".csv"]

    def parse(self, file_path: str) -> str:
        ext = file_path.rsplit(".", 1)[-1].lower()
        if ext == "csv":
            return self._parse_csv(file_path)
        return self._parse_excel(file_path)

    def _parse_csv(self, file_path: str) -> str:
        rows: list[str] = []
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(" | ".join(row))
        return "\n".join(rows)

    def _parse_excel(self, file_path: str) -> str:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel parsing: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_text: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"## Sheet: {sheet_name}\n")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    all_text.append(" | ".join(cells))

        wb.close()
        return "\n".join(all_text)
